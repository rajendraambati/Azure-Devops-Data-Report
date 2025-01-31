import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import streamlit as st
from io import BytesIO
import tempfile
import zipfile
import shutil

# Streamlit App Title
st.title("Excel File Processor")

# Step 1: Upload Folder
uploaded_folder = st.file_uploader("Upload a folder containing Excel files", type=["zip"], accept_multiple_files=False)

if uploaded_folder is not None:
    # Get the original filename without extension
    folder_name = os.path.splitext(uploaded_folder.name)[0]
    
    # Create a temporary directory using the uploaded file's name
    with tempfile.TemporaryDirectory() as temp_dir:
        extract_path = os.path.join(temp_dir, folder_name)
        os.makedirs(extract_path, exist_ok=True)
        
        # Extract the uploaded zip folder
        with zipfile.ZipFile(uploaded_folder, 'r') as zip_ref:
            zip_ref.extractall(extract_path)
        
        # Initialize dictionary to store employee details
        employee_details = {}

        # Function to extract iteration number
        def extract_iteration_number(iteration_name):
            match = re.search(r'ITR[_]?(\d+)', iteration_name)
            if match:
                return match.group(1)
            return None

        # Function to normalize iteration number
        def normalize_iteration_number(iteration_number):
            if iteration_number:
                return str(int(iteration_number))
            return None

        # Function to extract project name from iteration name
        def extract_project_name(iteration_name):
            if '\\' in iteration_name:
                return iteration_name.split('\\')[0]
            return iteration_name

        # Step 3: Loop through each Excel file in the folder
        for file in os.listdir(extract_path):
            if file.endswith('.xlsx') or file.endswith('.xls'):
                file_path = os.path.join(extract_path, file)
                data = pd.read_excel(file_path)

                # Check if required columns are present
                if 'Resource Name' in data.columns and 'Iteration Name' in data.columns and 'Project Name' in data.columns and 'Current Day work' in data.columns:
                    # Loop through each row and gather the required data
                    for _, row in data.iterrows():
                        employee_name = row['Resource Name']
                        iteration_name = row['Iteration Name']
                        project_name = row['Project Name']
                        current_day_work = row['Current Day work']

                        # Extract and normalize iteration number
                        iteration_number = extract_iteration_number(iteration_name)
                        new_iteration_number = normalize_iteration_number(iteration_number)

                        # Extract project name from iteration name
                        updated_project_name = extract_project_name(iteration_name)

                        if employee_name not in employee_details:
                            employee_details[employee_name] = {
                                'iterations': [],
                                'projects': set(),
                                'current_day_work': [],
                                'new_iteration_numbers': []
                            }

                        # Collect iterations, projects, current day work, and new iteration numbers
                        employee_details[employee_name]['iterations'].append(iteration_name)
                        employee_details[employee_name]['projects'].add(updated_project_name)
                        employee_details[employee_name]['current_day_work'].append(current_day_work)
                        if new_iteration_number:
                            employee_details[employee_name]['new_iteration_numbers'].append(new_iteration_number)

        # Step 4: Prepare data for the new Excel file
        rows = []
        for employee_name, details in employee_details.items():
            unique_new_iteration_numbers = list(set(details['new_iteration_numbers']))
            
            for new_iter_num in unique_new_iteration_numbers:
                filtered_iterations = [iter_name for iter_name, new_num in zip(details['iterations'], details['new_iteration_numbers']) if new_num == new_iter_num]
                filtered_current_day_work = [work for work, new_num in zip(details['current_day_work'], details['new_iteration_numbers']) if new_num == new_iter_num]
                
                current_day_work_sum = sum(filtered_current_day_work)
                unique_filtered_iterations = list(set(filtered_iterations))
                
                row = {
                    'Month Name': folder_name,
                    'Resource Name': employee_name,
                    'Iteration Name': ', '.join(unique_filtered_iterations),
                    'Project Name': ', '.join(details['projects']),
                    'Current Day work': current_day_work_sum,
                    'New Iteration Number': new_iter_num
                }
                rows.append(row)

        # Create DataFrame with the new rows
        unique_data = pd.DataFrame(rows)

        # Step 5: Create Excel file in memory
        output_buffer = BytesIO()
        unique_data.to_excel(output_buffer, index=False)
        
        # Step 6: Load workbook from buffer and highlight employee names
        output_buffer.seek(0)
        wb = load_workbook(output_buffer)
        ws = wb.active
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # Apply green highlight to the 'Resource Name' column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.fill = green_fill

        # Save the highlighted Excel file to buffer
        output_buffer = BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        # Step 7: Provide download link for the generated Excel file
        st.download_button(
            label="Download Processed File",
            data=output_buffer,
            file_name=f"processed_{folder_name}_details.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )